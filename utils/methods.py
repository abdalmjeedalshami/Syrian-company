import tkinter as tk
from tkinter import ttk, StringVar
from tkinter import messagebox, filedialog

# utils/submit_method.py

import os
import shutil
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from models.form_data import FormData


# Make sure to define these somewhere or pass them in
EXCEL_FILE = "data.xlsx"
PROFILE_FOLDER = "profiles"


def submit_data(form_data: FormData):
    # required_fields = [
    #     form_data.first_name, form_data.last_name, form_data.national_number,
    #     form_data.city, form_data.profile_path, form_data.gender, form_data.date
    # ]

    # if any(not field for field in required_fields):
    #     messagebox.showwarning("معلومات مفقودة", "الرجاء تعبئة كل الحقول")
    #     return

    try:
        wb = load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(
            [
                "الاسم",
                "اسم الأب",
                "الكنية",
                "اسم الأم",
                "الجنس",
                "تاريخ الولادة",
                "محل الولادة",
                "قيد النفوس",
                "الرقم الوطني",
                "العنوان المفصّل",
                "صورة وثيقة شخصية",
                "المستوى العلمي",
                "الاختصاص",
                "المهنة التي يتقنها",
                "المهنة الحالية",
                "الطول",
                "الوزن",
                "لون العينين",
                "لون الشعر",
                "لون البشرة",
                "علامات فارقة",
                "الوضع الاجتماعي",
                "عدد الأولاد",
                "الأولاد",
                "رقم الهاتف",
                "التواصل الاجتماعي",
                "الجنسية",
                "الجنسية الثانية"
            ]
        )
    else:
        ws = wb.active

    children_str = ", ".join(
        [f"{name} ({gender})" for name, gender in form_data.children]
    )

    ws.append(
        [
            form_data.first_name,
            form_data.father_name,
            form_data.last_name,
            form_data.mother_name,
            form_data.gender,
            form_data.date,
            form_data.city,
            form_data.residential_registration,
            form_data.national_number,
            form_data.full_address,
            form_data.profile_path,
            form_data.education,
            form_data.academic_specialization,
            form_data.job,
            form_data.current_job,
            form_data.height,
            form_data.weight,
            form_data.eye_color,
            form_data.hair_color,
            form_data.skin_color,
            form_data.unique_features,
            form_data.social_status,
            form_data.children_number,
            children_str,
            form_data.phone_number,
            form_data.social_media,
            form_data.nationality,
        ]
    )

    row = ws.max_row

    def set_hyperlink(col, label, path):
        cell = ws.cell(row=row, column=col)
        cell.value = label
        cell.hyperlink = path.replace("\\", "/")
        cell.font = Font(color="0000FF", underline="single")

    user_folder = os.path.join(PROFILE_FOLDER, form_data.first_name)
    os.makedirs(user_folder, exist_ok=True)

    filename = os.path.basename(form_data.profile_path)
    saved_path = os.path.join(user_folder, filename)
    shutil.copy(form_data.profile_path, saved_path)

    set_hyperlink(11, "عرض الصورة", saved_path)

    wb.save(EXCEL_FILE)
    messagebox.showinfo("تمت العملية بنجاح", "تم إضافة بيانات المستفيد")


def file_picker(on_file_picked, title="اختر ملف", first_name=""):
    if not first_name:
        messagebox.showwarning(
            "معلومات ناقصة", "يرجى إدخال الاسم أولاً قبل اختيار الصورة"
        )
        return

    filepath = filedialog.askopenfilename(
        title=title, filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp")]
    )

    if filepath:
        on_file_picked(filepath)


def generate_entries(children_frame, children_number_entry, children_entries):
    # Clear previous entries
    for widget in children_frame.winfo_children():
        widget.destroy()

    try:
        num_children = int(children_number_entry.get())  # Get the entered number
        for i in range(num_children):
            child_frame = ttk.Frame(children_frame)
            child_frame.pack(pady=10)
            ttk.Label(child_frame, text=f"اسم الولد {i + 1}").pack(
                side="right", padx=5, pady=2
            )
            entry = ttk.Entry(child_frame)
            entry.pack(side="right", padx=5, pady=2)

            selected_gender = StringVar()
            ttk.Radiobutton(
                child_frame, text="ذكر", variable=selected_gender, value="ذكر"
            ).pack(side="right", padx=5)
            ttk.Radiobutton(
                child_frame, text="أنثى", variable=selected_gender, value="انثى"
            ).pack(side="right", padx=5)

            children_entries.append(
                (entry, selected_gender)
            )  # Store both name entry & gender selection

    except ValueError:
        ttk.Label(children_frame, text="يرجى إدخال رقم صحيح!", foreground="red").pack(
            side="top", pady=5
        )


def create_labeled_entry(parent, label_text):
    frame = ttk.Frame(parent)
    frame.pack(pady=10, anchor="center")
    ttk.Label(frame, text=label_text).pack(side="right", padx=5)
    entry = ttk.Entry(frame)
    entry.pack(side="right", padx=5)
    return entry


def create_labeled_combobox(parent, label_text, values):
    frame = ttk.Frame(parent)
    frame.pack(pady=10)
    ttk.Label(frame, text=label_text).pack(side="right", padx=5)
    combobox = ttk.Combobox(frame, values=values, state="readonly")
    combobox.pack(side="right", padx=5)
    return combobox


def create_labeled_spinbox(parent, label_text, from_, to):
    frame = ttk.Frame(parent)
    frame.pack(pady=10)
    ttk.Label(frame, text=label_text).pack(side="right", padx=5)
    spinbox = ttk.Spinbox(frame, from_=from_, to=to, width=5)
    spinbox.pack(side="right", padx=5)
    return spinbox


def create_labeled_button(parent, label_text, button_text, command):
    frame = ttk.Frame(parent)
    frame.pack(pady=10)
    ttk.Label(frame, text=label_text).pack(side="right", padx=5)
    ttk.Button(frame, text=button_text, command=command).pack(side="right", padx=5)

def create_radio_group(parent, label_text, options, variable=None):
    frame = ttk.Frame(parent)
    frame.pack(pady=10)

    if variable is None:
        variable = tk.StringVar()

    ttk.Label(frame, text=label_text).pack(side="right", padx=5)

    for option in options:
        ttk.Radiobutton(frame, text=option, variable=variable, value=option).pack(
            side="right", padx=5
        )

    return variable

def create_date_picker(parent, label_text="تاريخ الولادة"):
    frame = ttk.Frame(parent)
    frame.pack(pady=10)

    ttk.Label(frame, text=label_text).pack(side="right", padx=5)

    # Day
    ttk.Label(frame, text="اليوم").pack(side="right")
    day_spinbox = ttk.Spinbox(frame, from_=1, to=31, width=5)
    day_spinbox.pack(side="right", padx=5)

    # Month
    ttk.Label(frame, text="الشهر").pack(side="right")
    month_spinbox = ttk.Spinbox(frame, from_=1, to=12, width=5)
    month_spinbox.pack(side="right", padx=5)

    # Year
    ttk.Label(frame, text="السنة").pack(side="right")
    year_spinbox = ttk.Spinbox(frame, from_=1900, to=2100, width=5)
    year_spinbox.pack(side="right", padx=5)

    return day_spinbox, month_spinbox, year_spinbox

def check_selection(event, combo, entry):
    """Shows an entry box if 'Other' is selected."""
    if combo.get() == "(غير)":
        entry.pack(side="left", padx=10)  # Show entry field
        entry.focus()  # Focus on entry field
    else:
        entry.pack_forget()  # Hide entry field

def clear_entries(entries):
    """Clears all provided entry fields."""
    for entry in entries:
        entry.delete(0, "end")

def reset_variables(variables):
    """Resets all provided variable fields."""
    for var in variables:
        var.set("")

def create_combobox_with_optional_entry(parent, label_text, options, other_label="(غير ذلك)"):
    frame = ttk.Frame(parent)
    frame.pack(pady=10)

    ttk.Label(frame, text=label_text).pack(side="right", padx=5)

    var = tk.StringVar()
    combobox = ttk.Combobox(
        frame,
        textvariable=var,
        values=options,
        state="readonly",
    )
    combobox.pack(side="right", padx=5)

    entry_var = tk.StringVar()
    entry = ttk.Entry(frame, textvariable=entry_var)
    entry.pack(side="right", padx=5)
    entry.pack_forget()

    def on_select(event):
        if var.get() == other_label:
            entry.pack(side="right", padx=5)
        else:
            entry.pack_forget()

    combobox.bind("<<ComboboxSelected>>", on_select)

    def get_value():
        return entry_var.get() if var.get() == other_label else var.get()

    return get_value, var, entry_var

def get_full_date(day, month, year):
    
    return f"{day}/{month}/{year}"

def clear_inputs_in_frame(frame):
    for child in frame.winfo_children():
        if isinstance(child, tk.Entry) or isinstance(child, ttk.Entry):
            child.delete(0, tk.END)
        elif isinstance(child, ttk.Combobox):
            child.set("")
        elif isinstance(child, tk.Text):
            child.delete("1.0", tk.END)
        elif hasattr(child, "winfo_children"):  # for nested frames
            clear_inputs_in_frame(child)

    # Clear all StringVar, IntVar, etc. attached to children
    for var in frame.children.values():
        if isinstance(var, (tk.StringVar, tk.IntVar, tk.BooleanVar)):
            var.set("")
