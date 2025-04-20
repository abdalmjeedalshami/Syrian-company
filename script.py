from tkinter import *
from tkinter import ttk, messagebox
from openpyxl import load_workbook, Workbook
import customtkinter as ctk


EXCEL_FILE = 'data.xlsx'  # Ensure this file exists
cities = ["دمشق", "حلب", "اللاذقية", "حمص", "طرطوس"]


def submit_data():
    firstName = first_name_entry.get()
    lastName = last_name_entry.get()
    national_number = national_number_entry.get()
    city = city_entry.get()

    if not firstName or not lastName or not national_number or not city:
        messagebox.showwarning("معلومات مفقودة", "الرجاء تعبئة كل الحقول")
        return

    try:
        wb = load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(["الاسم الأول", "الاسم الأخير", "الرقم الوطني", "المدينة"])  # Add headers
    else:
        ws = wb.active

    ws.append([firstName, lastName, national_number, city])
    wb.save(EXCEL_FILE)

    first_name_entry.delete(0, END)
    last_name_entry.delete(0, END)
    national_number_entry.delete(0, END)
    city_entry.delete(0, END)
    city_entry.set("")

    messagebox.showinfo("تمت العملية بنجاح", "تم إضافة بيانات المستفيد")

# Create GUI
root = Tk()
root.iconbitmap("app_icon.ico")
root.title("استبيان إضافة مستفيد")

# Maximize window (Windows-specific)
root.wm_state('zoomed')  

# Make root window responsive
root.columnconfigure(0, weight=1)
root.rowconfigure(0, weight=1)

# Create fram
frm = ttk.Frame(root, padding=20)
frm.grid()

# Entry fields
first_name_entry = ttk.Entry(frm)
last_name_entry = ttk.Entry(frm)
national_number_entry = ttk.Entry(frm)
city_entry = ttk.Combobox(frm, values=cities, state="readonly")

# Labels
ttk.Label(frm, text="الاسم الأول").grid(row=0, column=1, padx=10, pady=5)
ttk.Label(frm, text="الاسم الأخير").grid(row=1, column=1, padx=10, pady=5)
ttk.Label(frm, text="الرقم الوطني").grid(row=2, column=1, padx=10, pady=5)
ttk.Label(frm, text="المدينة").grid(row=3, column=1, padx=10, pady=5)

# Entry fields pos
first_name_entry.grid(row=0, column=0, padx=10, pady=8)
last_name_entry.grid(row=1, column=0, padx=10, pady=8)
national_number_entry.grid(row=2, column=0, padx=10, pady=8)
city_entry.grid(row=3, column=0, padx=10, pady=8)

# Buttons
ttk.Button(frm, text="إضافة", command=submit_data).grid(row=4, column=0, columnspan=2, pady=10)
# submit_button = ctk.CTkButton(frm, text="إضافة", command=submit_data).grid(row=5, column=0)

root.mainloop()
