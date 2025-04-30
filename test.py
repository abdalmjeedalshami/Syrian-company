from tkinter import *
from tkinter import ttk, messagebox, font, filedialog
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
import customtkinter as ctk
import os
import shutil


EXCEL_FILE = 'data.xlsx'  # Ensure this file exists
PROFILE_FOLDER = 'images/profile'
cities = ["دمشق", "حلب", "اللاذقية", "حمص", "طرطوس"]

# Function to generate entries based on user input
def generate_entries():
    # Clear previous entries
    for widget in children_frame.winfo_children():
        widget.destroy()

    try:
        num_children = int(children_number_entry.get())  # Get the entered number
        children_entries.clear()  # Clear previous entries

        for i in range(num_children):
            child_frame = ttk.Frame(children_frame)
            child_frame.pack(pady=10)

            ttk.Label(child_frame, text=f"اسم الولد {i + 1}").pack(side="right", padx=5, pady=2)
            entry = ttk.Entry(child_frame)
            entry.pack(side="right", padx=5, pady=2)

            selected_gender = StringVar()
            ttk.Radiobutton(child_frame, text="ذكر", variable=selected_gender, value="ذكر").pack(side="right", padx=5)
            ttk.Radiobutton(child_frame, text="أنثى", variable=selected_gender, value="أنثى").pack(side="right", padx=5)

            children_entries.append((entry, selected_gender))  # Store both name entry & gender selection

    except ValueError:
        ttk.Label(children_frame, text="يرجى إدخال رقم صحيح!", foreground="red").pack(side="top", pady=5)


# Update scroll region when adding new widgets
def update_scroll_region(event):
    canvas.configure(scrollregion=canvas.bbox("all"))

# Recalculate centering after resize
def resize_centering(event):
    canvas.coords(content_window, (canvas.winfo_width() // 2, 0))

def get_full_date(day, month, year):
    return f"{day}-{month}-{year}"

def set_profile_path(path):
    global profile_path
    profile_path = path

def file_picker(on_file_picked, title='اختر ملف'):
    first_name = first_name_entry.get()

    if not first_name:
        messagebox.showwarning("معلومات ناقصة", "يرجى إدخال الاسم أولاً قبل اختيار الصورة")
        return

    # Choose image
    filepath = filedialog.askopenfilename(
         title=title,
        filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp")]
    )

    if filepath:
         on_file_picked(filepath)

def submit_data(first_name, last_name, national_number, city, profile_path, gender, date, children):
    
    if not first_name or not last_name or not national_number or not city or not profile_path or not gender or not date:
        messagebox.showwarning("معلومات مفقودة", "الرجاء تعبئة كل الحقول")
        return
    
    try:
        wb = load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(["الاسم الأول", "الاسم الأخير", "الرقم الوطني", "المدينة", "الصورة الشخصية", "الجنس", "تاريخ الولادة"])  # Add headers
    else:
        ws = wb.active

    ws.append([first_name, last_name, national_number, city, "", gender, date])
    row = ws.max_row

    # Helper to add hyperlink to a cell
    def set_hyperlink(col, label, path):
            cell = ws.cell(row=row, column=col)
            cell.value = label
            cell.hyperlink = path.replace("\\", "/")
            cell.font = Font(color="0000FF", underline="single")

    # Create folder based on first name
    user_folder = os.path.join(PROFILE_FOLDER, first_name)
    os.makedirs(user_folder, exist_ok=True)

    # Copy image to user folder
    filename = os.path.basename(profile_path)
    saved_path = os.path.join(user_folder, filename)
    shutil.copy(profile_path, saved_path)

    # Set hyperlinks
    set_hyperlink(5, "عرض الصورة", saved_path)

    wb.save(EXCEL_FILE)

    first_name_entry.delete(0, END)
    last_name_entry.delete(0, END)
    national_number_entry.delete(0, END)
    city_entry.set("")
    selected_gender.set("")
    selected_day.set("")
    selected_month.set("")
    selected_year.set("")

    messagebox.showinfo("تمت العملية بنجاح", "تم إضافة بيانات المستفيد")


# Create GUI
root = Tk()
root.iconbitmap("app_icon.ico")
root.title("استبيان إضافة مستفيد")
root.geometry("500x500")

# Maximize window (Windows-specific)
root.wm_state('zoomed')  

# Custom fonts
normal_font = font.Font(family="Segoe UI")
bold_font = font.Font(family="Segoe UI", weight="bold")

# Create a frame for the scrollbar and canvas
scroll_frame = ttk.Frame(root)
scroll_frame.pack(fill="both", expand=True)

# Create canvas to hold the scrollable content
canvas = Canvas(scroll_frame)
canvas.pack(side="left", fill="both", expand=True)

# Add scrollbar
scrollbar = ttk.Scrollbar(scroll_frame, orient="vertical", command=canvas.yview)
scrollbar.pack(side="right", fill="y")

# Create a frame inside the canvas to hold the actual content
content_frame = ttk.Frame(canvas)

# Add window to canvas, using `winfo_width()` for dynamic centering
canvas.create_window((canvas.winfo_width() // 2, 0), window=content_frame, anchor="n")

# Configure scrollbar
canvas.configure(yscrollcommand=scrollbar.set)

# Create main frame
# frm = ttk.Frame(root, padding=20)
# frm.pack(fill="both", expand=True, padx=20, pady=20)

# Title
ttk.Label(content_frame, text="استمارة (معتقل - مختفي - ناجي - شاهد)", font=bold_font).pack(pady=10)

# Entry fields
first_name_frame = ttk.Frame(content_frame)
first_name_frame.pack(pady=10)
ttk.Label(first_name_frame, text="الاسم الأول").pack(side="right", padx=5)
first_name_entry = ttk.Entry(first_name_frame)
first_name_entry.pack(side="right", padx=5)

father_name_frame = ttk.Frame(content_frame)
father_name_frame.pack(pady=10)
ttk.Label(father_name_frame, text="الاسم الأب").pack(side="right", padx=5)
father_name_entry = ttk.Entry(father_name_frame)
father_name_entry.pack(side="right", padx=5)

last_name_frame = ttk.Frame(content_frame)
last_name_frame.pack(pady=10)
ttk.Label(last_name_frame, text="الكنية").pack(side="right", padx=5)
last_name_entry = ttk.Entry(last_name_frame)
last_name_entry.pack(side="right", padx=5)

mother_name_frame = ttk.Frame(content_frame)
mother_name_frame.pack(pady=10)
ttk.Label(mother_name_frame, text="اسم الأم").pack(side="right", padx=5)
mother_name_entry = ttk.Entry(mother_name_frame)
mother_name_entry.pack(side="right", padx=5)

gender_frame = ttk.Frame(content_frame)
gender_frame.pack(pady=10)
selected_gender = StringVar()
ttk.Label(gender_frame, text="الجنس").pack(side="right", padx=5)
ttk.Radiobutton(gender_frame, text="ذكر", variable=selected_gender, value="ذكر").pack(side="right", padx=5)
ttk.Radiobutton(gender_frame, text="أنثى", variable=selected_gender, value="أنثى").pack(side="right", padx=5)

date_frame = ttk.Frame(content_frame)
date_frame.pack(pady=10)
ttk.Label(date_frame, text="تاريخ الولادة").pack(side="right", padx=5)
ttk.Label(date_frame, text="اليوم").pack(side="right")
selected_day = ttk.Spinbox(date_frame, from_=1, to=31, width=5)
selected_day.pack(side="right", padx=5)
ttk.Label(date_frame, text="الشهر").pack(side="right")
selected_month = ttk.Spinbox(date_frame, from_=1, to=12, width=5)
selected_month.pack(side="right", padx=5)
ttk.Label(date_frame, text="السنة").pack(side="right")
selected_year = ttk.Spinbox(date_frame, width=5)
selected_year.pack(side="right", padx=5)

city_frame = ttk.Frame(content_frame)
city_frame.pack(pady=10)
ttk.Label(city_frame, text="محل الولادة").pack(side="right", padx=5)
city_entry = ttk.Combobox(city_frame, values=["دمشق", "حلب", "اللاذقية"], state="readonly")
city_entry.pack(side="right", padx=5)

residential_registration_frame = ttk.Frame(content_frame)
residential_registration_frame.pack(pady=10)
ttk.Label(residential_registration_frame, text="قيد النفوس").pack(side="right", padx=5)
residential_registration_entry = ttk.Entry(residential_registration_frame)
residential_registration_entry.pack(side="right", padx=5)

national_number_frame = ttk.Frame(content_frame)
national_number_frame.pack(pady=10)
ttk.Label(national_number_frame, text="الرقم الوطني").pack(side="right", padx=5)
national_number_entry = ttk.Entry(national_number_frame)
national_number_entry.pack(side="right", padx=5)

full_address_frame = ttk.Frame(content_frame)
full_address_frame.pack(pady=10)
ttk.Label(full_address_frame, text="العنوان المفصّل").pack(side="right", padx=5)
full_address_entry = ttk.Entry(full_address_frame)
full_address_entry.pack(side="right", padx=5)

profile_frame = ttk.Frame(content_frame)
profile_frame.pack(pady=10)
profile_path = None
ttk.Label(profile_frame, text="الصورة الشخصية").pack(side="right", padx=5)
ttk.Button(profile_frame, text="اختر الصورة الشخصية", command=lambda: file_picker(set_profile_path, title="اختر صورة شخصية")).pack(side="right", padx=5)

education_frame = ttk.Frame(content_frame)
education_frame.pack(pady=10)
ttk.Label(education_frame, text="المستوى العلمي").pack(side="right", padx=5)
education_entry = ttk.Combobox(education_frame, values=["أمّي", "ابتدائي", "إعدادي", "ثانوي", "معهد", "جامعة", "ماجستر", "دكتور"], state="readonly")
education_entry.pack(side="right", padx=5)

academic_specialization_frame = ttk.Frame(content_frame)
academic_specialization_frame.pack(pady=10)
ttk.Label(academic_specialization_frame, text="الاختصاص").pack(side="right", padx=5)
academic_specialization_entry = ttk.Entry(academic_specialization_frame)
academic_specialization_entry.pack(side="right", padx=5)

job_frame = ttk.Frame(content_frame)
job_frame.pack(pady=10)
ttk.Label(job_frame, text="المهنة التي يتقنها").pack(side="right", padx=5)
job_entry = ttk.Entry(job_frame)
job_entry.pack(side="right", padx=5)

current_job_frame = ttk.Frame(content_frame)
current_job_frame.pack(pady=10)
ttk.Label(current_job_frame, text="المهنة الحالية").pack(side="right", padx=5)
current_job_entry = ttk.Entry(current_job_frame)
current_job_entry.pack(side="right", padx=5)

height_frame = ttk.Frame(content_frame)
height_frame.pack(pady=10)
ttk.Label(height_frame, text="الطول").pack(side="right", padx=5)
ttk.Label(height_frame, text="سم").pack(side="right")
selected_height = ttk.Spinbox(height_frame, width=5)
selected_height.pack(side="right", padx=5)

weight_frame = ttk.Frame(content_frame)
weight_frame.pack(pady=10)
ttk.Label(weight_frame, text="الوزن").pack(side="right", padx=5)
ttk.Label(weight_frame, text="كغ").pack(side="right")
selected_weight = ttk.Spinbox(weight_frame, width=5)
selected_weight.pack(side="right", padx=5)

eye_color_frame = ttk.Frame(content_frame)
eye_color_frame.pack(pady=10)
ttk.Label(eye_color_frame, text="لون العينين").pack(side="right", padx=5)
eye_color_entry = ttk.Combobox(eye_color_frame, values=["أسود", "بني", "أزرق", "أخضر", "عسلي"], state="readonly")
eye_color_entry.pack(side="right", padx=5)

hair_color_frame = ttk.Frame(content_frame)
hair_color_frame.pack(pady=10)
ttk.Label(hair_color_frame, text="لون الشعر").pack(side="right", padx=5)
hair_color_entry = ttk.Combobox(hair_color_frame, values=["أسود", "بني", "أشقر", "أحمر"], state="readonly")
hair_color_entry.pack(side="right", padx=5)

skin_color_frame = ttk.Frame(content_frame)
skin_color_frame.pack(pady=10)
ttk.Label(skin_color_frame, text="لون البشرة").pack(side="right", padx=5)
skin_color_entry = ttk.Combobox(skin_color_frame, values=["أبيض", "حنطي", "أسمر", "أسود"], state="readonly")
skin_color_entry.pack(side="right", padx=5)

unique_features_frame = ttk.Frame(content_frame)
unique_features_frame.pack(pady=10)
ttk.Label(unique_features_frame, text="علامات فارقة").pack(side="right", padx=5)
unique_features_entry = ttk.Entry(unique_features_frame)
unique_features_entry.pack(side="right", padx=5)

skin_color_frame = ttk.Frame(content_frame)
skin_color_frame.pack(pady=10)
ttk.Label(skin_color_frame, text=" ؟ ؟الوضع الاجتماعي").pack(side="right", padx=5)
skin_color_entry = ttk.Combobox(skin_color_frame, values=["أبيض", "حنطي", "أسمر", "أسود"], state="readonly")
skin_color_entry.pack(side="right", padx=5)

# Entry for number of children
children_frame = ttk.Frame(content_frame)
children_frame.pack(pady=10)
ttk.Label(children_frame, text="عدد الأطفال").pack(side="right", pady=5)
children_number_entry = ttk.Entry(children_frame)
children_number_entry.pack(side="right", pady=5)

# Button to generate child entries
ttk.Button(children_frame, text="إنشاء حقول الأولاد", command=generate_entries).pack(side="right", pady=5)

# Frame to hold generated child entry fields
children_frame = ttk.Frame(content_frame)
children_frame.pack(pady=10)

# List to store child entry widgets
children_entries = []


# Buttons
add_btn_frame = ttk.Frame(content_frame)
add_btn_frame.pack(pady=10)
ttk.Button(
    add_btn_frame,
    text="إضافة", 
    command=lambda: submit_data(
        first_name=first_name_entry.get(),
        last_name=last_name_entry.get(),
        national_number=national_number_entry.get(),
        city=city_entry.get(),
        profile_path=profile_path,
        gender=selected_gender.get(),
        date=get_full_date(selected_day.get(), selected_month.get(), selected_year.get()))
).pack(pady=10)
# submit_button = ctk.CTkButton(frm, text="إضافة", command=submit_data).grid(row=5, column=0)

content_window = canvas.create_window((canvas.winfo_width() // 2, 0), window=content_frame, anchor="n")
canvas.bind("<Configure>", resize_centering)  # Recenter on window resize
content_frame.bind("<Configure>", update_scroll_region)
root.mainloop()
